import openpyxl

# Crear un nuevo libro de Excel
workbook = openpyxl.load_workbook()

# Seleccionar la hoja activa
sheet = workbook.active
sheet.title = "Lista_de_Supermercados"

# Agregar encabezados
sheet['A1'] = "Supermercado"
sheet['B1'] = "Producto"
sheet['C1'] = "Precio"

# Lista de supermercados con nombre de producto y precio
supermarkets = [
    ("Supermercado A", "Manzanas", 2.99),
    ("Supermercado A", "Peras", 3.49),
    ("Supermercado B", "Naranjas", 2.79),
    ("Supermercado B", "Kiwi", 3.29),
    ("Supermecado B", "Huevos", 5.75),
    ("Supermecado A", "Pan Tostado", 12.05),
    ("Supermecado A", "Coca Cola", 4.20)
]

# Agregar datos a la hoja
for row, data in enumerate(supermarkets, start=2):
    sheet.cell(row=row, column=1, value=data[0])
    sheet.cell(row=row, column=2, value=data[1])
    sheet.cell(row=row, column=3, value=data[2])

# Guardar el archivo Excel
workbook.save("lista_supermercados.xlsx")

# Cerrar el archivo Excel
workbook.close()
